import os
import time
from flask import Flask, render_template, request, send_file, redirect, url_for, flash
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'segredo')  # Melhor para produção

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Configurações para produção
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        start_time = time.time()
        
        excel = request.files.get('excel')
        datas = request.files.get('datas')
        coluna = request.form.get('coluna', 'F')
        cor = request.form.get('cor', '#FFFF00')

        if not excel or not datas:
            flash('Por favor, envie ambos os arquivos.')
            return redirect(request.url)

        # Verificar tamanho dos arquivos
        excel.seek(0, 2)  # Ir para o final
        excel_size = excel.tell()
        excel.seek(0)  # Voltar ao início
        
        if excel_size > 10 * 1024 * 1024:  # 10MB
            flash('Arquivo Excel muito grande. Máximo 10MB.')
            return redirect(request.url)

        excel_filename = secure_filename(excel.filename)
        datas_filename = secure_filename(datas.filename)
        excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)
        datas_path = os.path.join(UPLOAD_FOLDER, datas_filename)
        
        try:
            excel.save(excel_path)
            datas.save(datas_path)
        except Exception as e:
            flash(f'Erro ao salvar arquivos: {e}')
            return redirect(request.url)

        # Processar arquivo de datas
        try:
            with open(datas_path, 'r', encoding='utf-8') as f:
                datas_monitoradas = [
                    datetime.strptime(linha.strip(), "%d/%m/%Y").date()
                    for linha in f if linha.strip()
                ]
        except Exception as e:
            flash(f"Erro ao ler o arquivo de datas: {e}")
            return redirect(request.url)

        # Carregar planilha com timeout
        try:
            wb = load_workbook(excel_path, read_only=True)  # Modo read-only para melhor performance
            aba = wb.active
        except Exception as e:
            flash(f"Erro ao abrir o arquivo Excel: {e}")
            return redirect(request.url)

        # Converter cor hex para formato aceito pelo openpyxl (sem #)
        cor_hex = cor.lstrip('#')
        preenchimento = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
        total_destacadas = 0
        
        # Lista para armazenar ocorrências encontradas
        ocorrencias = []

        # Processar linhas com verificação de timeout
        max_rows = min(aba.max_row, 10000)  # Limitar a 10.000 linhas para evitar timeout
        
        for linha in range(2, max_rows + 1):
            # Verificar timeout a cada 1000 linhas
            if linha % 1000 == 0:
                if time.time() - start_time > 25:  # 25 segundos de timeout
                    flash('Processamento muito lento. Tente com uma planilha menor.')
                    return redirect(request.url)
            
            try:
                celula = aba[f"{coluna}{linha}"]
                valor = celula.value
                if isinstance(valor, datetime):
                    if valor.date() in datas_monitoradas:
                        total_destacadas += 1
                        
                        # Adicionar à lista de ocorrências
                        ocorrencias.append({
                            'linha': linha,
                            'data': valor.strftime("%d/%m/%Y"),
                            'valor_celula': str(valor),
                            'coluna': coluna
                        })
            except Exception as e:
                continue  # Pular linhas com erro

        # Fechar planilha original
        wb.close()

        # Criar relatório de ocorrências
        relatorio_path = os.path.join(UPLOAD_FOLDER, 'relatorio_ocorrencias.xlsx')
        wb_relatorio = Workbook()
        ws_relatorio = wb_relatorio.active
        ws_relatorio.title = "Relatório de Ocorrências"

        # Cabeçalhos do relatório
        ws_relatorio['A1'] = "Linha"
        ws_relatorio['B1'] = "Data Encontrada"
        ws_relatorio['C1'] = "Valor da Célula"
        ws_relatorio['D1'] = "Coluna"
        ws_relatorio['E1'] = "Status"

        # Estilo para cabeçalhos
        header_font = Font(bold=True)
        for cell in ws_relatorio[1]:
            cell.font = header_font

        # Preencher dados das ocorrências
        for i, ocorrencia in enumerate(ocorrencias, start=2):
            ws_relatorio[f'A{i}'] = ocorrencia['linha']
            ws_relatorio[f'B{i}'] = ocorrencia['data']
            ws_relatorio[f'C{i}'] = ocorrencia['valor_celula']
            ws_relatorio[f'D{i}'] = ocorrencia['coluna']
            ws_relatorio[f'E{i}'] = "ENCONTRADA"

        # Adicionar resumo
        linha_resumo = len(ocorrencias) + 3
        ws_relatorio[f'A{linha_resumo}'] = "RESUMO:"
        ws_relatorio[f'A{linha_resumo + 1}'] = f"Total de ocorrências encontradas: {total_destacadas}"
        ws_relatorio[f'A{linha_resumo + 2}'] = f"Datas monitoradas: {len(datas_monitoradas)}"
        ws_relatorio[f'A{linha_resumo + 3}'] = f"Datas monitoradas: {', '.join([d.strftime('%d/%m/%Y') for d in datas_monitoradas])}"
        ws_relatorio[f'A{linha_resumo + 4}'] = f"Linhas processadas: {max_rows - 1}"
        ws_relatorio[f'A{linha_resumo + 5}'] = f"Tempo de processamento: {time.time() - start_time:.2f} segundos"

        # Ajustar largura das colunas
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_relatorio.column_dimensions[col].width = 15

        wb_relatorio.save(relatorio_path)
        wb_relatorio.close()

        flash(f'Processamento concluído! {total_destacadas} ocorrências encontradas na coluna {coluna}.')
        
        # Retornar relatório
        if total_destacadas > 0:
            return send_file(relatorio_path, as_attachment=True, download_name='relatorio_ocorrencias.xlsx')
        else:
            flash('Nenhuma ocorrência encontrada.')
            return redirect(request.url)

    return render_template('index.html')

# Configuração para produção (Render)
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)